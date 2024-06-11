# Author Serkan Güldal 2021.09.19
from itertools import filterfalse
import os
from pathlib import Path

from sklearn import metrics
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
plt.style.use('fivethirtyeight')

from sklearn.ensemble import AdaBoostClassifier
from imblearn.metrics import *

from collections import Counter
from sklearn.model_selection import *
from sklearn.metrics import *

from pandas import DataFrame
from pandas import read_csv
from numpy import mean
from matplotlib import pyplot


import multiprocessing as mp
import time
import openpyxl

debug = False # It provides more detailed output for debugging and extented analysis.

def data(inputFile): # Data importer function
    file = open(os.path.dirname(__file__) + '/../datasets/' + inputFile)
    df=pd.read_csv(file)
    global X
    X = df.values[:,0:-1]
    global y
    y = df.values[:,-1]
    return(X, y)

print_feature_importance_once = True
def ml(X, y, r): # Machine learning approach
    global print_feature_importance_once

    X_train, X_test, y_train, y_test = train_test_split(X, y, train_size=0.6, random_state=42, shuffle=True, stratify=y)
    c = AdaBoostClassifier(random_state=42, algorithm="SAMME")
    c.fit(X_train, y_train)
    y_pred = c.predict(X_test)

    if debug:
        print("Round ", r)

        if print_feature_importance_once:
            # get importance
            importance = c.feature_importances_
            # summarize feature importance
            for i,v in enumerate(importance):
                print('Feature: %0d, Score: %.5f' % (i,v))
            # plot feature importance
            pyplot.bar([x for x in range(len(importance))], importance)
            pyplot.show()
            print_feature_importance_once = False

    # Measurements
    Accuracy = accuracy_score(y_test, y_pred)    
    AreaUnderROCcurve = roc_auc_score(y_test, c.predict_proba(X_test)[:, 1])
    
    fpr, tpr, thresholds = metrics.roc_curve(y_test, y_pred, pos_label=0)
    AreaUndercurve0 = metrics.auc(fpr, tpr)

    fpr, tpr, thresholds = metrics.roc_curve(y_test, y_pred, pos_label=1)
    AreaUndercurve1 = metrics.auc(fpr, tpr)

    Recall = recall_score(y_test, y_pred, average='macro')
    Precision = precision_score(y_test, y_pred, average='macro')
    f1score = f1_score(y_test, y_pred, average='macro')
    specificity = specificity_score(y_test, y_pred, average='macro')
    sensitivity = sensitivity_score(y_test, y_pred, average='macro')
    geometric = geometric_mean_score(y_test, y_pred, average='macro')

    return(Accuracy, AreaUnderROCcurve, AreaUndercurve0, AreaUndercurve1, Recall, Precision, f1score, specificity, sensitivity, geometric)

def moving_average(x):
    global ave
    ave = []
    for i in range(len(x)):
        m = mean(x[0:i+1])
        ave.append(m)
    return ave


if __name__ == '__main__':
    
    input_raw = 'yeast3_label_class.csv' # Original filename needs to be updated!!!!
    input = input_raw + '_10_SMOTE.csv' # Used modification names needs to be added if there is any.
    print(input)

    # Load the existing Excel file or create a new one if it doesn't exist
    excel_file_path = os.path.dirname(__file__) + '/../ML_Results/' + input_raw + '_AdaBoost.xlsx'
    if os.path.exists(excel_file_path):
        wb = openpyxl.load_workbook(excel_file_path)
    else:
        # Create a new Workbook object
        wb = openpyxl.Workbook()
        wb.save(excel_file_path)

    sheet_names = wb.sheetnames

    # Set one of the existing sheets as the active sheet
    sheet = wb[sheet_names[0]]

    # Iterate through the rows and check if the input value exists
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value == input:
            input_row = row
            print(f"\nPrevious results are in line {row}. They will be overwritten!!!")
            break
    else:
        # Input value not found, write to the next empty row
        input_row = sheet.max_row + 1

    sheet.cell(row=1, column=1,  value='Input File')
    sheet.cell(row=1, column=2,  value='Accuracy')
    sheet.cell(row=1, column=3,  value='Area Under ROC curve')
    sheet.cell(row=1, column=4,  value='Area Under the Curve 0')
    sheet.cell(row=1, column=5,  value='Area Under the Curve 1')
    sheet.cell(row=1, column=6,  value='Recall')
    sheet.cell(row=1, column=7,  value='Precision')
    sheet.cell(row=1, column=8,  value='F1 Score')
    sheet.cell(row=1, column=9,  value='Specificity')
    sheet.cell(row=1, column=10, value='Sensitivity')
    sheet.cell(row=1, column=11, value='Geometric Mean')
    sheet.cell(row=1, column=12, value='Arithmetic Mean')
    sheet.cell(row=1, column=13, value='Total time')


# REMOVE after all the following methods are introduced in the resampling methods (oversampling/undersampling)

    # for file in [
    #     # '',
    #     # '_ADASYN.csv',
    #     # '_AllKNN.csv',
    #     # '_BorderlineSMOTE.csv',
    #     # '_ClusterCentroids.csv',
    #     # '_CondensedNearestNeighbour.csv',
    #     # '_EditedNearestNeighbours.csv',
    #     # '_InstanceHardnessThreshold.csv',
    #     # '_NearMiss.csv',
    #     # '_NeighbourhoodCleaningRule.csv',
    #     # '_OneSidedSelection.csv',
    #     # '_RandomOverSampler.csv',
    #     # '_RandomUnderSampler.csv',
    #     # '_RepeatedEditedNearestNeighbours.csv',
    #     # '_SMOTE.csv',
    #     # '_SMOTEENN.csv',
    #     # '_SMOTEN.csv',
    #     # '_SMOTETomek.csv',
    #     # '_SVMSMOTE.csv',
    #     # '_TomekLinks.csv',
    #     # #Our methods
    #     # '_AGG_WA_CD.csv',
    #     # '_GM_WA_CD.csv',
    #     # '_Heinz.csv',
    #     # '_Weighted.csv',
    #     # '_OverSampling_Arithmetic_Random.csv',

    #     ]:

    data(input) # Data importer function
    ts = time.time()    
    
    accuracies = []
    a = []
    mean_acc_old = 0
    tolerance = 0.0001
    num_iterations = 5000  # Change this value to adjust the maxiumum number of iterations
    for r in range(1, num_iterations):
        
        result = ml(X, y, r)
        a.append(result)

        acc = result[0]
        accuracies.append(acc)
        
        # Calculate the mean accuracy
        mean_acc = sum(accuracies) / len(accuracies)

        if abs(mean_acc - mean_acc_old) < tolerance:
            print("Iteration is converged at step",r,"...\n")
            break

        mean_acc_old = mean_acc

        # Check if this is the last iteration
        if r == num_iterations-1:
            print("Reached the highest value of r. ML result may not be correct?!\n")



    score = np.array(a)
    acc = score[:,0]
    aucroc = score[:,1]
    auc0 = score[:,2]
    auc1 = score[:,3]
    rc = score[:,4]
    pre = score[:,5]
    f = score[:,6]
    sp = score[:,7]
    sen = score[:,8]
    geo = score[:,9]
    aveALL = mean([ mean(acc), mean(aucroc), mean(auc0), mean(auc1), mean(rc), mean(pre), mean(f), mean(sp), mean(sen)])
    duration = time.time() - ts

    data_type = input

    sheet.cell(row=input_row, column=1,  value=input)
    sheet.cell(row=input_row, column=2,  value=mean(acc))
    sheet.cell(row=input_row, column=3,  value=mean(aucroc))
    sheet.cell(row=input_row, column=4,  value=mean(auc0))
    sheet.cell(row=input_row, column=5,  value=mean(auc1))
    sheet.cell(row=input_row, column=6,  value=mean(rc))
    sheet.cell(row=input_row, column=7,  value=mean(pre))
    sheet.cell(row=input_row, column=8,  value=mean(f))
    sheet.cell(row=input_row, column=9,  value=mean(sp))
    sheet.cell(row=input_row, column=10, value=mean(sen))
    sheet.cell(row=input_row, column=11, value=mean(geo))
    sheet.cell(row=input_row, column=12, value=mean(aveALL))
    sheet.cell(row=input_row, column=13, value=mean(duration))
    wb.save(excel_file_path)

    if debug:
        print(input + ' is completed. Here is the summary.')
        print("Accuracy:",mean(acc))
        print("Area Under ROC curve:", mean(aucroc))
        print("Area Under the Curve 0:", mean(auc0))
        print("Area Under the Curve 1:", mean(auc1))
        print("Recall:", mean(rc))
        print("Precision:", mean(pre))
        print("F1 Score:", mean(f))
        print("Specificity:", mean(sp))
        print("Sensitivity:", mean(sen))
        print("Geometric Mean:", mean(geo))
        print("Arithmetic Mean:", mean(aveALL))        
        print('Time in parallel:', duration)

        ###### Writing to files ########

        # with open('RF_' + dataname + '_' + 'scores.csv', 'w') as filehandle:
        #     for listitem in score:
        #         filehandle.write('%s\n' % listitem)

        debug_folder_path = os.path.dirname(__file__) + '/../ML_Results/' + input + '/AdaBoost_debug'

        if not os.path.exists(debug_folder_path):
            # Create the folder
            Path(debug_folder_path).mkdir(parents=True, exist_ok=True)


        prename = debug_folder_path + '/ADaBoost_' + data_type + '_'

        with open(prename + 'All_Score.csv', 'w') as filehandle:
            for listitem in score:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'Accuracy.csv', 'w') as filehandle:
            for listitem in acc:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'AUCROC.csv', 'w') as filehandle:
            for listitem in aucroc:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'AUC_0.csv', 'w') as filehandle:
            for listitem in auc0:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'AUC_1.csv', 'w') as filehandle:
            for listitem in auc1:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'Recall.csv', 'w') as filehandle:
            for listitem in rc:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'Precision.csv', 'w') as filehandle:
            for listitem in pre:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'F1_Score.csv', 'w') as filehandle:
            for listitem in f:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'Specificity.csv', 'w') as filehandle:
            for listitem in sp:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'Sensitivity.csv', 'w') as filehandle:
            for listitem in sen:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'Geometric.csv', 'w') as filehandle:
            for listitem in geo:
                filehandle.write('%s\n' % listitem)

        with open(prename + 'summary.txt', 'w') as filehandle:
            filehandle.write('              Accuracy: %s' % mean(acc))
            filehandle.write('\n  Area Under ROC curve: %s' % mean(aucroc))
            filehandle.write('\nArea Under the Curve 0: %s' % mean(auc0))
            filehandle.write('\nArea Under the Curve 1: %s' % mean(auc1))
            filehandle.write('\n                Recall: %s' % mean(rc))
            filehandle.write('\n             Precision: %s' % mean(pre))
            filehandle.write('\n              F1 Score: %s' % mean(f))
            filehandle.write('\n           Specificity: %s' % mean(sp))
            filehandle.write('\n           Sensitivity: %s' % mean(sen))
            filehandle.write('\n        Geometric Mean: %s' % mean(geo))
            filehandle.write('\n       Arithmetic Mean: %s' % aveALL)
            filehandle.write('\n            Total time: %s' % duration)

        with open(prename + 'Time.csv', 'w') as filehandle:
            filehandle.write('Total time is %s' % duration)

        moving_average(acc)
        with open(prename + 'Accuracy_Convergence.csv', 'w') as filehandle:
            for listitem in ave:
                filehandle.write('%s\n' % listitem)

        time.sleep(5)

        pyplot.plot(acc)
        pyplot.plot(ave)
        pyplot.savefig(prename + 'accuracy.png')
        pyplot.show()

        # pyplot.plot(acc)
        # pyplot.plot([mean(acc) for x in range(len(acc))])
        # pyplot.show()